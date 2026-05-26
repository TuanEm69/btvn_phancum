import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sklearn.cluster import KMeans

# ==============================================================================
# 1. ĐỌC DỮ LIỆU VÀ SÀNG LỌC CỘT TRỐNG TOÀN BỘ (ALL NAN)
# ==============================================================================
try:
    df_lvc = pd.read_excel("DiemTongHop.xlsx", header=None)
except FileNotFoundError:
    print("Không tìm thấy file 'DiemTongHop.xlsx'. Vui lòng đặt file cùng thư mục với đoạn code này.")
    exit()

# Trích xuất thông tin định danh sinh viên ban đầu ở dòng 1 và dòng 2
mssv_row = df_lvc.iloc[1, 3:].values
tensv_row = df_lvc.iloc[2, 3:].values

# Trích xuất ma trận điểm số sinh viên (Hàng 4 đến hàng 55)
students_df = df_lvc.iloc[4:56, 3:].copy()
students_df = students_df.apply(pd.to_numeric, errors='coerce')

# Tìm ra các cột trống TOÀN BỘ (Tất cả 52 ô điểm của cột đó đều là NaN)
cols_to_drop = students_df.columns[students_df.isna().all(axis=0)]

# Tiến hành loại bỏ các cột trống toàn bộ ra khỏi ma trận và mảng thông tin định danh
students_df_filtered = students_df.drop(columns=cols_to_drop)
mssv_filtered = np.delete(mssv_row, [c - 3 for c in cols_to_drop])
tensv_filtered = np.delete(tensv_row, [c - 3 for c in cols_to_drop])

# Đối với các cột còn lại (có vài ô trống cục bộ), gán mặc định bằng 0.0 theo yêu cầu
students_df_final = students_df_filtered.fillna(0.0)

# Tính điểm trung bình cộng (GPA) cho từng sinh viên hợp lệ còn lại
gpa_scores = students_df_final.mean(axis=0)

# Tạo bảng dữ liệu đặc trưng sinh viên sau khi sàng lọc
features_filtered = pd.DataFrame({
    'MSSV': mssv_filtered,
    'Ten_SV': tensv_filtered,
    'GPA': gpa_scores.values
})

# ==============================================================================
# 2. ÁP DỤNG THUẬT TOÁN K-MEANS PHÂN THÀNH 3 CỤM THEO GPA MỚI
# ==============================================================================
X = features_filtered[['GPA']].values
kmeans = KMeans(n_clusters=3, random_state=42, n_init=10)
features_filtered['Cluster_Label'] = kmeans.fit_predict(X)

# Sắp xếp tên cụm theo thứ tự điểm trung bình GPA tăng dần
cluster_means = features_filtered.groupby('Cluster_Label')['GPA'].mean().sort_values()
cluster_names = {
    cluster_means.index[0]: "Trung bình / Yếu",
    cluster_means.index[1]: "Khá",
    cluster_means.index[2]: "Giỏi"
}
features_filtered['Ten_Nhom_Cum'] = features_filtered['Cluster_Label'].map(cluster_names)

print("Thống kê số lượng sinh viên sau khi áp dụng quy tắc lọc mới:")
print(features_filtered['Ten_Nhom_Cum'].value_counts())

# ==============================================================================
# 3. TRỰC QUAN HÓA KẾT QUẢ (SCATTER PLOT & PIE CHART)
# ==============================================================================
# Đồ thị 1: Biểu đồ phân cụm điểm tọa độ
plt.figure(figsize=(11, 4))
colors_map = {"Giỏi": '#2E7D32', "Khá": '#FBC02D', "Trung bình / Yếu": '#C62828'}

for group_name in ["Giỏi", "Khá", "Trung bình / Yếu"]:
    sub_df = features_filtered[features_filtered['Ten_Nhom_Cum'] == group_name]
    y_jitter = np.random.normal(0, 0.05, size=len(sub_df))
    plt.scatter(sub_df['GPA'], y_jitter, c=colors_map[group_name], label=group_name, s=100, edgecolors='black', alpha=0.85)

# Tính ranh giới phân cụm
boundary_1 = (features_filtered[features_filtered['Ten_Nhom_Cum'] == "Trung bình / Yếu"]['GPA'].max() + 
              features_filtered[features_filtered['Ten_Nhom_Cum'] == "Khá"]['GPA'].min()) / 2
boundary_2 = (features_filtered[features_filtered['Ten_Nhom_Cum'] == "Khá"]['GPA'].max() + 
              features_filtered[features_filtered['Ten_Nhom_Cum'] == "Giỏi"]['GPA'].min()) / 2

plt.axvline(x=boundary_1, color='#757575', linestyle='--', linewidth=1.5)
plt.axvline(x=boundary_2, color='#757575', linestyle='--', linewidth=1.5)

plt.title('Biểu đồ phân bố cụm Sinh viên (Chỉ loại bỏ cột trống toàn bộ)', fontsize=13, fontweight='bold', pad=15)
plt.xlabel('Điểm GPA hệ 4', fontsize=11, fontweight='bold')
plt.xlim(1.0, 4.0)
plt.ylim(-0.3, 0.3)
plt.gca().get_yaxis().set_visible(False)
plt.grid(axis='x', linestyle=':', alpha=0.6)
plt.legend(loc='upper left', title="Phân nhóm cụm", title_fontproperties={'weight': 'bold'})
plt.tight_layout()
plt.savefig('bieu_do_phan_cum_chuan_cot.png', dpi=300)
plt.close()

# Đồ thị 2: Biểu đồ tròn tỷ lệ phần trăm
labels = ["Giỏi", "Khá", "Trung bình / Yếu"]
counts = [features_filtered[features_filtered['Ten_Nhom_Cum'] == l].shape[0] for l in labels]

plt.figure(figsize=(8, 7))
plt.pie(counts, labels=labels, autopct='%1.1f%%', startangle=140, colors=['#2E7D32', '#FBC02D', '#C62828'],
        textprops={'fontsize': 11, 'fontweight': 'bold'})
plt.title('Biểu đồ tỷ lệ các cụm học lực (Chỉ loại bỏ cột trống toàn bộ)', fontsize=13, fontweight='bold', pad=20)
plt.axis('equal')
plt.tight_layout()
plt.savefig('bieu_do_tron_chuan_cot.png', dpi=300)
plt.close()

# ==============================================================================
# 4. ĐỊNH DẠNG VÀ XUẤT FILE EXCEL CÓ DÒNG TỔNG SỐ NGƯỜI
# ==============================================================================
# Sắp xếp danh sách theo độ ưu tiên cụm hiển thị và điểm GPA giảm dần
features_filtered['Cluster_Priority'] = features_filtered['Ten_Nhom_Cum'].map({"Giỏi": 0, "Khá": 1, "Trung bình / Yếu": 2})
features_filtered = features_filtered.sort_values(by=['Cluster_Priority', 'GPA'], ascending=[True, False]).reset_index(drop=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Phân Cụm GPA Chuẩn Cột"
ws.views.sheetView[0].showGridLines = True

HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")

FONT_TITLE = Font(name="Segoe UI", size=15, bold=True, color="1F497D")
FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Segoe UI", size=10)
FONT_BODY_BOLD = Font(name="Segoe UI", size=10, bold=True)
FONT_TOTAL = Font(name="Segoe UI", size=10, bold=True, color="1F497D")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                     top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

ws.merge_cells("A2:E2")
ws["A2"] = "DANH SÁCH PHÂN CỤM GPA (CHỈ LOẠI BỎ CỘT TRỐNG TOÀN BỘ)"
ws["A2"].font = FONT_TITLE
ws["A2"].alignment = ALIGN_CENTER

headers_columns = ["STT", "MSSV", "Tên Sinh Viên", "Điểm GPA lớp học", "Nhóm Phân Cụm (K=3)"]
for c_idx, h in enumerate(headers_columns, 1):
    cell = ws.cell(row=4, column=c_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = HEADER_FILL
    cell.alignment = ALIGN_CENTER

current_row = 5
stt_counter = 1

for group_name in ["Giỏi", "Khá", "Trung bình / Yếu"]:
    group_df = features_filtered[features_filtered['Ten_Nhom_Cum'] == group_name]
    start_sub_row = current_row
    
    for idx, row in group_df.iterrows():
        ws.cell(row=current_row, column=1, value=stt_counter).alignment = ALIGN_CENTER
        ws.cell(row=current_row, column=2, value=row['MSSV']).alignment = ALIGN_CENTER
        ws.cell(row=current_row, column=3, value=row['Ten_SV']).alignment = ALIGN_LEFT
        
        c_gpa = ws.cell(row=current_row, column=4, value=round(row['GPA'], 2))
        c_gpa.number_format = '0.00'
        c_gpa.alignment = ALIGN_CENTER
        
        c_name = ws.cell(row=current_row, column=5, value=row['Ten_Nhom_Cum'])
        c_name.alignment = ALIGN_CENTER
        
        fill = ZEBRA_FILL if stt_counter % 2 == 1 else WHITE_FILL
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.fill = fill
            cell.font = FONT_BODY if col != 5 else FONT_BODY_BOLD
            
        current_row += 1
        stt_counter += 1
        
    end_sub_row = current_row - 1
    
    # Tạo hàng tổng kết số lượng tự động cho từng cụm
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    sum_label_cell = ws.cell(row=current_row, column=1, value=f"Tổng số sinh viên cụm {group_name}:")
    sum_label_cell.font = FONT_TOTAL
    sum_label_cell.alignment = ALIGN_RIGHT
    
    sum_val_cell = ws.cell(row=current_row, column=4, value=f"=COUNTA(B{start_sub_row}:B{end_sub_row})")
    sum_val_cell.font = FONT_TOTAL
    sum_val_cell.alignment = ALIGN_CENTER
    
    ws.cell(row=current_row, column=5, value="")
    
    for col in range(1, 6):
        cell = ws.cell(row=current_row, column=col)
        cell.border = Border(top=Side(style='thin', color='1F497D'), bottom=Side(style='medium', color='1F497D'),
                             left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'))
        cell.fill = TOTAL_FILL
        
    current_row += 2

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 28

output_excel = "Ket_Qua_KMeans_GPA_Chuan_Cot.xlsx"
wb.save(output_excel)
print(f"-> Đã hoàn thành xử lý phân cụm dựa trên cột hợp lệ và xuất file thành công!")